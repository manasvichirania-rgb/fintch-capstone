"""
build_workbook.py

Part A: builds merchant_workbook.xlsx from ledger.csv and merchants.csv.

Sheets:
    Merchants        - raw merchant reference data
    Transactions_View - every ledger row + VLOOKUP-derived merchant fields,
                        an HLOOKUP-derived fee tier, and the nested IF/AND
                        "High-Value Merchant Day" classification
    Fee_Tiers        - horizontal MDR-style fee-tier reference table (HLOOKUP source)
    Daily_Merchant   - one row per (merchant_id, date) with a live SUMIFS daily total
                        (this is the "pivot" the classification rule reads from)
    Pivot_Summary    - amount+count by merchant_id/status, and a count-vs-unique-days
                        comparison per merchant

The workbook uses live SUMIFS/COUNTIFS summary tables for the daily and merchant
summaries. These formulas recalculate when the transaction data changes.

Run:
    python build_workbook.py
    python /mnt/skills/public/xlsx/scripts/recalc.py merchant_workbook.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_PATH = "merchant_workbook.xlsx"

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E2E40", end_color="2E2E40", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=13)
NOTE_FONT = Font(name="Arial", italic=True, size=9, color="555555")


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    ledger = pd.read_csv("ledger.csv", parse_dates=["transaction_time"])
    merchants = pd.read_csv("merchants.csv")

    n_txn = len(ledger)
    n_merch = len(merchants)
    last_txn_row = n_txn + 1     # header is row 1
    last_merch_row = n_merch + 1

    wb = Workbook()

    # Workbook notes and assumptions
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme["A1"] = "Merchant Workbook — Paytm Payments & Fraud Analytics (Part 1A)"
    ws_readme["A1"].font = TITLE_FONT
    notes = [
        "",
        "Sheets:",
        "  Merchants          - reference table of the 40 merchants (id, name, category, region)",
        "  Transactions_View  - all 547 ledger transactions with VLOOKUP-derived merchant fields,",
        "                       an HLOOKUP-derived payment-method fee tier, and the nested",
        "                       IF/AND 'High-Value Merchant Day' classification",
        "  Fee_Tiers          - horizontal MDR-style fee-tier reference table (HLOOKUP source)",
        "  Daily_Merchant     - one row per (merchant_id, date) with a live SUMIFS daily total;",
        "                       this is the table the classification rule reads its daily total from",
        "  Pivot_Summary      - total amount_inr + transaction count by merchant_id and status,",
        "                       plus a transaction-count-vs-unique-days comparison per merchant",
        "",
        "Design decisions / documented assumptions:",
        "",
        "1. Fee tiers (HLOOKUP, Fee_Tiers sheet): UPI 0.30%, Wallet 0.60%, Card 1.80%,",
        "   Netbanking 0.90%. These are ILLUSTRATIVE assumptions for this assignment,",
        "   not real Paytm MDR rates.",
        "",
        "2. 'High-Value Merchant Day' classification rule (nested IF/AND, Transactions_View",
        "   column O): a transaction is labeled 'High-Value Merchant Day' when its merchant's",
        "   TOTAL transaction amount on that CALENDAR DAY exceeds INR 5,000 AND the merchant's",
        "   region is NOT 'East'. The daily total is pulled from the Daily_Merchant table",
        "   (column N, via SUMIFS), matching the assignment's requirement that the daily total",
        "   come from a pivot/summary table rather than being recomputed inline per row.",
        "",
        "3. Pivot-table substitution: openpyxl cannot reliably write a native Excel PivotTable",
        "   object from scratch. Daily_Merchant and Pivot_Summary are built instead as live",
        "   SUMIFS/COUNTIFS cross-tabs referencing Transactions_View directly - they recalculate",
        "   automatically if the transaction data changes, the same way a refreshed pivot table",
        "   would, and are NOT hardcoded result values.",
        "",
        "4. All VLOOKUP formulas use a fixed absolute range (Merchants!$A$2:$D$41) and are",
        "   wrapped in IFERROR(...,\"Merchant not found\") to handle any unmatched merchant_id.",
        "",
        "5. All monetary values are in INR.",
    ]
    for i, line in enumerate(notes, start=2):
        ws_readme.cell(row=i, column=1, value=line).font = BODY_FONT
    ws_readme.column_dimensions["A"].width = 105

    # Merchant reference data
    ws_m = wb.create_sheet("Merchants")
    ws_m.append(["merchant_id", "merchant_name", "category", "region"])
    for _, r in merchants.iterrows():
        ws_m.append([int(r.merchant_id), r.merchant_name, r.category, r.region])
    style_header_row(ws_m, 1, 4)
    for row in ws_m.iter_rows(min_row=2, max_row=last_merch_row, max_col=4):
        for cell in row:
            cell.font = BODY_FONT
    autosize(ws_m, [12, 16, 16, 10])
    ws_m.freeze_panes = "A2"

    # Fee lookup table used by HLOOKUP
    ws_f = wb.create_sheet("Fee_Tiers")
    ws_f["A1"] = "Payment Method (illustrative MDR fee tiers for this assignment)"
    ws_f["A1"].font = Font(name="Arial", bold=True, italic=True, size=10)
    ws_f["A2"] = "Fee % of transaction amount"
    ws_f["A2"].font = BODY_FONT
    methods = ["UPI", "Wallet", "Card", "Netbanking"]
    fee_pcts = [0.003, 0.006, 0.018, 0.009]
    for i, (m, p) in enumerate(zip(methods, fee_pcts)):
        col = get_column_letter(2 + i)  # B, C, D, E
        ws_f[f"{col}1"] = m
        ws_f[f"{col}1"].font = HEADER_FONT
        ws_f[f"{col}1"].fill = HEADER_FILL
        ws_f[f"{col}1"].alignment = Alignment(horizontal="center")
        ws_f[f"{col}2"] = p
        ws_f[f"{col}2"].number_format = "0.00%"
        ws_f[f"{col}2"].font = BODY_FONT
    ws_f["A4"] = ("NOTE: these fee percentages are illustrative assumptions chosen for this "
                  "assignment and do not represent real Paytm merchant discount rates (MDR).")
    ws_f["A4"].font = NOTE_FONT
    autosize(ws_f, [16, 10, 10, 10, 12])

    # Daily merchant totals
    ledger["txn_date"] = ledger["transaction_time"].dt.date
    daily_pairs = (
        ledger.groupby(["merchant_id", "txn_date"])
        .size()
        .reset_index(name="n")
        .sort_values(["merchant_id", "txn_date"])
    )
    merch_region_lookup = merchants.set_index("merchant_id")["region"].to_dict()

    ws_d = wb.create_sheet("Daily_Merchant")
    ws_d.append(["merchant_id", "date", "daily_total_amount_inr", "region"])
    style_header_row(ws_d, 1, 4)
    dm_first_data_row = 2
    for i, (_, r) in enumerate(daily_pairs.iterrows(), start=dm_first_data_row):
        mid = int(r.merchant_id)
        ws_d.cell(row=i, column=1, value=mid).font = BODY_FONT
        date_cell = ws_d.cell(row=i, column=2, value=r.txn_date)
        date_cell.number_format = "yyyy-mm-dd"
        date_cell.font = BODY_FONT
        # SUMIFS using merchant and transaction date
        formula = (f"=SUMIFS(Transactions_View!$E$2:$E${last_txn_row},"
                   f"Transactions_View!$C$2:$C${last_txn_row},A{i},"
                   f"Transactions_View!$L$2:$L${last_txn_row},B{i})")
        total_cell = ws_d.cell(row=i, column=3, value=formula)
        total_cell.number_format = "#,##0"
        total_cell.font = BODY_FONT
        region_formula = f'=IFERROR(VLOOKUP(A{i},Merchants!$A$2:$D$41,4,FALSE),"Merchant not found")'
        ws_d.cell(row=i, column=4, value=region_formula).font = BODY_FONT
    dm_last_row = dm_first_data_row + len(daily_pairs) - 1
    autosize(ws_d, [12, 14, 22, 10])
    ws_d.freeze_panes = "A2"

    # Transaction view
    ws_t = wb.create_sheet("Transactions_View")
    headers = ["transaction_id", "user_id", "merchant_id", "transaction_time", "amount_inr",
               "payment_method", "status", "risk_score", "merchant_name (VLOOKUP)",
               "category (VLOOKUP)", "region (VLOOKUP)", "txn_date", "fee_tier_pct (HLOOKUP)",
               "merchant_daily_total_inr (from Daily_Merchant pivot)",
               "classification (nested IF/AND)"]
    ws_t.append(headers)
    style_header_row(ws_t, 1, len(headers))

    for i, (_, r) in enumerate(ledger.iterrows(), start=2):
        ws_t.cell(row=i, column=1, value=r.transaction_id).font = BODY_FONT
        ws_t.cell(row=i, column=2, value=int(r.user_id)).font = BODY_FONT
        ws_t.cell(row=i, column=3, value=int(r.merchant_id)).font = BODY_FONT
        dt_cell = ws_t.cell(row=i, column=4, value=r.transaction_time.to_pydatetime())
        dt_cell.number_format = "yyyy-mm-dd hh:mm"
        dt_cell.font = BODY_FONT
        amt_cell = ws_t.cell(row=i, column=5, value=int(r.amount_inr))
        amt_cell.number_format = "#,##0"
        amt_cell.font = BODY_FONT
        ws_t.cell(row=i, column=6, value=r.payment_method).font = BODY_FONT
        ws_t.cell(row=i, column=7, value=r.status).font = BODY_FONT
        ws_t.cell(row=i, column=8, value=int(r.risk_score)).font = BODY_FONT

        # Merchant fields via VLOOKUP
        f_name = f'=IFERROR(VLOOKUP(C{i},Merchants!$A$2:$D$41,2,FALSE),"Merchant not found")'
        ws_t.cell(row=i, column=9, value=f_name).font = BODY_FONT
        # Category lookup
        f_cat = f'=IFERROR(VLOOKUP(C{i},Merchants!$A$2:$D$41,3,FALSE),"Merchant not found")'
        ws_t.cell(row=i, column=10, value=f_cat).font = BODY_FONT
        # Region lookup
        f_region = f'=IFERROR(VLOOKUP(C{i},Merchants!$A$2:$D$41,4,FALSE),"Merchant not found")'
        ws_t.cell(row=i, column=11, value=f_region).font = BODY_FONT
        # L: date-only (drops the time component so it matches Daily_Merchant's date column)
        f_date = f"=DATE(YEAR(D{i}),MONTH(D{i}),DAY(D{i}))"
        date_only_cell = ws_t.cell(row=i, column=12, value=f_date)
        date_only_cell.number_format = "yyyy-mm-dd"
        date_only_cell.font = BODY_FONT
        # M: fee tier % via HLOOKUP against Fee_Tiers!$B$1:$E$2
        f_hlookup = f'=IFERROR(HLOOKUP(F{i},Fee_Tiers!$B$1:$E$2,2,FALSE),"Fee tier not found")'
        fee_cell = ws_t.cell(row=i, column=13, value=f_hlookup)
        fee_cell.number_format = "0.00%"
        fee_cell.font = BODY_FONT
        # N: merchant's daily total, pulled from the Daily_Merchant pivot via SUMIFS
        f_daily = (f"=SUMIFS(Daily_Merchant!$C${dm_first_data_row}:$C${dm_last_row},"
                   f"Daily_Merchant!$A${dm_first_data_row}:$A${dm_last_row},C{i},"
                   f"Daily_Merchant!$B${dm_first_data_row}:$B${dm_last_row},L{i})")
        daily_cell = ws_t.cell(row=i, column=14, value=f_daily)
        daily_cell.number_format = "#,##0"
        daily_cell.font = BODY_FONT
        # High-value day classification -> "High-Value Merchant Day" when
        #    merchant's daily total > 5000 AND region <> "East"
        f_class = f'=IF(AND(N{i}>5000,K{i}<>"East"),"High-Value Merchant Day","")'
        ws_t.cell(row=i, column=15, value=f_class).font = BODY_FONT

    autosize(ws_t, [14, 9, 11, 17, 11, 14, 11, 10, 16, 15, 10, 12, 13, 20, 24])
    ws_t.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # Sheet: Pivot_Summary
    # ------------------------------------------------------------------
    ws_p = wb.create_sheet("Pivot_Summary")
    ws_p["A1"] = "Section 1 - Total amount_inr and transaction count by merchant_id and status"
    ws_p["A1"].font = Font(name="Arial", bold=True, size=11)

    hdr_row = 2
    section1_headers = ["merchant_id", "merchant_name",
                         "captured_amount_inr", "captured_count",
                         "failed_amount_inr", "failed_count",
                         "chargeback_amount_inr", "chargeback_count",
                         "total_amount_inr", "total_count"]
    for c, h in enumerate(section1_headers, start=1):
        ws_p.cell(row=hdr_row, column=c, value=h)
    style_header_row(ws_p, hdr_row, len(section1_headers))

    statuses = ["captured", "failed", "chargeback"]
    s1_first_data_row = hdr_row + 1
    for i, mid in enumerate(sorted(merchants["merchant_id"].tolist()), start=s1_first_data_row):
        ws_p.cell(row=i, column=1, value=mid).font = BODY_FONT
        f_mname = f'=IFERROR(VLOOKUP(A{i},Merchants!$A$2:$D$41,2,FALSE),"Merchant not found")'
        ws_p.cell(row=i, column=2, value=f_mname).font = BODY_FONT
        col = 3
        for status in statuses:
            f_amt = (f'=SUMIFS(Transactions_View!$E$2:$E${last_txn_row},'
                     f'Transactions_View!$C$2:$C${last_txn_row},A{i},'
                     f'Transactions_View!$G$2:$G${last_txn_row},"{status}")')
            amt_cell = ws_p.cell(row=i, column=col, value=f_amt)
            amt_cell.number_format = "#,##0"
            amt_cell.font = BODY_FONT
            f_cnt = (f'=COUNTIFS(Transactions_View!$C$2:$C${last_txn_row},A{i},'
                     f'Transactions_View!$G$2:$G${last_txn_row},"{status}")')
            ws_p.cell(row=i, column=col + 1, value=f_cnt).font = BODY_FONT
            col += 2
        f_total_amt = f"=C{i}+E{i}+G{i}"
        tot_amt_cell = ws_p.cell(row=i, column=9, value=f_total_amt)
        tot_amt_cell.number_format = "#,##0"
        tot_amt_cell.font = BODY_FONT
        f_total_cnt = f"=D{i}+F{i}+H{i}"
        ws_p.cell(row=i, column=10, value=f_total_cnt).font = BODY_FONT
    s1_last_row = s1_first_data_row + n_merch - 1

    section2_title_row = s1_last_row + 3
    ws_p.cell(row=section2_title_row, column=1,
              value="Section 2 - Transaction count vs. unique transaction days, per merchant")
    ws_p.cell(row=section2_title_row, column=1).font = Font(name="Arial", bold=True, size=11)

    s2_hdr_row = section2_title_row + 1
    section2_headers = ["merchant_id", "merchant_name", "total_transaction_count",
                         "unique_transaction_days", "avg_txns_per_active_day"]
    for c, h in enumerate(section2_headers, start=1):
        ws_p.cell(row=s2_hdr_row, column=c, value=h)
    style_header_row(ws_p, s2_hdr_row, len(section2_headers))

    s2_first_data_row = s2_hdr_row + 1
    for i, mid in enumerate(sorted(merchants["merchant_id"].tolist()), start=s2_first_data_row):
        ws_p.cell(row=i, column=1, value=mid).font = BODY_FONT
        f_mname = f'=IFERROR(VLOOKUP(A{i},Merchants!$A$2:$D$41,2,FALSE),"Merchant not found")'
        ws_p.cell(row=i, column=2, value=f_mname).font = BODY_FONT
        # total transaction count for this merchant
        f_total_cnt = f'=COUNTIF(Transactions_View!$C$2:$C${last_txn_row},A{i})'
        ws_p.cell(row=i, column=3, value=f_total_cnt).font = BODY_FONT
        # unique days transacted = count of rows for this merchant in Daily_Merchant
        # (Daily_Merchant has exactly one row per merchant/date pair with >=1 txn)
        f_unique_days = f'=COUNTIF(Daily_Merchant!$A${dm_first_data_row}:$A${dm_last_row},A{i})'
        ws_p.cell(row=i, column=4, value=f_unique_days).font = BODY_FONT
        f_avg = f'=IFERROR(C{i}/D{i},0)'
        avg_cell = ws_p.cell(row=i, column=5, value=f_avg)
        avg_cell.number_format = "0.00"
        avg_cell.font = BODY_FONT

    ws_p.column_dimensions["A"].width = 12
    ws_p.column_dimensions["B"].width = 16
    for col_letter in ["C", "D", "E", "F", "G", "H", "I", "J"]:
        ws_p.column_dimensions[col_letter].width = 18
    ws_p.freeze_panes = "A3"

    wb.save(OUT_PATH)
    print(f"Saved {OUT_PATH}")
    print(f"Transactions_View rows: {n_txn} (rows 2-{last_txn_row})")
    print(f"Daily_Merchant rows: {len(daily_pairs)} (rows {dm_first_data_row}-{dm_last_row})")
    print(f"Pivot_Summary section 1: rows {s1_first_data_row}-{s1_last_row}")
    print(f"Pivot_Summary section 2: rows {s2_first_data_row}-{s2_first_data_row + n_merch - 1}")


if __name__ == "__main__":
    main()
